import streamlit as st
import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import calendar

st.set_page_config(
    page_title="Mercado 8AM – Consumo por Mes",
    page_icon="🛒",
    layout="wide"
)

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    [data-testid="stAppViewContainer"] { background: #f8f9fc; }
    .block-container { padding-top: 1.5rem; }
    .metric-card {
        background: white;
        border-radius: 12px;
        padding: 1.2rem 1.5rem;
        box-shadow: 0 2px 8px rgba(0,0,0,.07);
        border-left: 4px solid #6366f1;
    }
    h1 { color: #1e1e2d; font-weight: 700; }
    .section-header {
        font-size: 1.1rem; font-weight: 600;
        color: #374151; margin: 1rem 0 .4rem;
    }
</style>
""", unsafe_allow_html=True)

EXCEL_PATH = "Mercado_8AM.xlsx"

# ── PARSER ────────────────────────────────────────────────────────────────────
@st.cache_data
def load_data(path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find blocks: a "product name" row is one where some cells are strings
        # and the rest are None, followed by a header row containing 'Abierto'.
        i = 0
        while i < len(rows):
            row = rows[i]
            # Look for product-name row
            product_cols = []
            for col_idx, cell in enumerate(row):
                if isinstance(cell, str) and cell.strip() and cell.strip() not in (
                    "Gramos", "Mililitros", "Cantidad", "Pedazos", "Pastillas",
                    "Gramos/día", "Gramos/dia", "Huevos/dia", "Pedazos/dia",
                    "Pastillas/día", "ml/dia", "pañales/dia"
                ):
                    product_cols.append((col_idx, cell.strip()))

            if product_cols and i + 1 < len(rows):
                header_row = rows[i + 1]
                # Check header_row contains 'Abierto'
                if any(isinstance(c, str) and "Abierto" in c for c in header_row if c):
                    # For each product in this row, determine its column offset
                    for p_idx, (p_col, p_name) in enumerate(product_cols):
                        # Skip sub-labels like 'Chocolate'
                        if p_name in ("Chocolate",):
                            continue
                        # Find the 'Abierto' column at or after p_col
                        abierto_col = None
                        for c_idx in range(p_col, len(header_row)):
                            if isinstance(header_row[c_idx], str) and "Abierto" in header_row[c_idx]:
                                abierto_col = c_idx
                                break
                        if abierto_col is None:
                            continue
                        qty_col = abierto_col - 1       # quantity / amount
                        acabado_col = abierto_col + 1   # end date

                        # Gather data rows until next blank or product row
                        j = i + 2
                        while j < len(rows):
                            drow = rows[j]
                            qty   = drow[qty_col]   if qty_col < len(drow) else None
                            fecha_abierto  = drow[abierto_col] if abierto_col < len(drow) else None
                            fecha_acabado  = drow[acabado_col] if acabado_col < len(drow) else None

                            # Stop if we hit another product-name block
                            if any(
                                isinstance(drow[k], str) and drow[k].strip()
                                and drow[k].strip() not in (
                                    "Gramos", "Mililitros", "Cantidad", "Pedazos",
                                    "Pastillas", "Gramos/día", "Gramos/dia",
                                    "Huevos/dia", "Pedazos/dia", "Pastillas/día",
                                    "ml/dia", "pañales/dia"
                                )
                                for k in range(len(drow)) if drow[k]
                            ):
                                break

                            if isinstance(fecha_abierto, datetime) and isinstance(qty, (int, float)):
                                end = fecha_acabado if isinstance(fecha_acabado, datetime) else None
                                records.append({
                                    "categoria": sheet_name,
                                    "producto": p_name,
                                    "cantidad": float(qty),
                                    "fecha_inicio": fecha_abierto,
                                    "fecha_fin": end,
                                    "mes_inicio": fecha_abierto.month,
                                    "anio_inicio": fecha_abierto.year,
                                    "mes_label": fecha_abierto.strftime("%Y-%m"),
                                })
                            j += 1
            i += 1

    wb.close()
    df = pd.DataFrame(records)
    return df


df = load_data(EXCEL_PATH)

# ── HEADER ────────────────────────────────────────────────────────────────────
st.title("🛒 Mercado 8AM – Consumo por Mes")
st.markdown("Visualización del historial de consumo del hogar por producto y categoría.")

# ── SIDEBAR FILTERS ───────────────────────────────────────────────────────────
with st.sidebar:
    st.header("Filtros")
    categorias = ["Todas"] + sorted(df["categoria"].unique().tolist())
    cat_sel = st.selectbox("Categoría", categorias)

    if cat_sel != "Todas":
        productos_disponibles = sorted(df[df["categoria"] == cat_sel]["producto"].unique())
    else:
        productos_disponibles = sorted(df["producto"].unique())

    prods_sel = st.multiselect(
        "Productos", productos_disponibles, default=productos_disponibles[:5]
    )

    meses_disponibles = sorted(df["mes_label"].unique())
    mes_rango = st.select_slider(
        "Rango de meses",
        options=meses_disponibles,
        value=(meses_disponibles[0], meses_disponibles[-1])
    )

# ── FILTER ────────────────────────────────────────────────────────────────────
dff = df.copy()
if cat_sel != "Todas":
    dff = dff[dff["categoria"] == cat_sel]
if prods_sel:
    dff = dff[dff["producto"].isin(prods_sel)]
dff = dff[(dff["mes_label"] >= mes_rango[0]) & (dff["mes_label"] <= mes_rango[1])]

# ── KPI ROW ───────────────────────────────────────────────────────────────────
total_consumos = len(dff)
total_productos = dff["producto"].nunique()
total_unidades  = dff["cantidad"].sum()
meses_cubiertos = dff["mes_label"].nunique()

c1, c2, c3, c4 = st.columns(4)
def kpi(col, label, value, unit=""):
    col.markdown(f"""
    <div class="metric-card">
        <div style="font-size:.8rem;color:#6b7280;">{label}</div>
        <div style="font-size:1.8rem;font-weight:700;color:#1e1e2d;">{value}</div>
        <div style="font-size:.75rem;color:#9ca3af;">{unit}</div>
    </div>""", unsafe_allow_html=True)

kpi(c1, "Registros en rango", total_consumos, "aperturas de producto")
kpi(c2, "Productos distintos", total_productos, "productos únicos")
kpi(c3, "Unidades totales", f"{total_unidades:,.0f}", "g / ml / uds")
kpi(c4, "Meses cubiertos", meses_cubiertos, "meses con datos")

st.divider()

# ── CHARTS ────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs(
    ["📅 Consumo mensual", "📦 Por producto", "🗂️ Por categoría", "📋 Datos crudos"]
)

# ── TAB 1: Monthly consumption bar chart ──────────────────────────────────────
with tab1:
    st.markdown('<div class="section-header">Unidades abiertas por mes y producto</div>', unsafe_allow_html=True)
    monthly = (
        dff.groupby(["mes_label", "producto"])["cantidad"]
        .sum()
        .reset_index()
        .sort_values("mes_label")
    )
    if monthly.empty:
        st.info("Sin datos para los filtros seleccionados.")
    else:
        fig = px.bar(
            monthly,
            x="mes_label", y="cantidad", color="producto",
            barmode="group",
            labels={"mes_label": "Mes", "cantidad": "Unidades", "producto": "Producto"},
            template="plotly_white",
            color_discrete_sequence=px.colors.qualitative.Pastel,
            height=420,
        )
        fig.update_layout(
            xaxis_title="Mes", yaxis_title="Cantidad (g / ml / uds)",
            legend_title="Producto",
            xaxis=dict(tickangle=-30),
            margin=dict(t=20, b=60),
        )
        st.plotly_chart(fig, use_container_width=True)

    # Line chart: total monthly
    st.markdown('<div class="section-header">Total de unidades por mes (todos los productos seleccionados)</div>', unsafe_allow_html=True)
    total_mes = (
        dff.groupby("mes_label")["cantidad"].sum().reset_index().sort_values("mes_label")
    )
    fig2 = px.line(
        total_mes, x="mes_label", y="cantidad",
        markers=True,
        labels={"mes_label": "Mes", "cantidad": "Unidades totales"},
        template="plotly_white",
        color_discrete_sequence=["#6366f1"],
        height=300,
    )
    fig2.update_traces(line_width=2.5, marker_size=7)
    fig2.update_layout(margin=dict(t=10, b=50), xaxis=dict(tickangle=-30))
    st.plotly_chart(fig2, use_container_width=True)

# ── TAB 2: Per-product treemap + boxplot ──────────────────────────────────────
with tab2:
    st.markdown('<div class="section-header">Consumo total por producto</div>', unsafe_allow_html=True)
    prod_total = (
        dff.groupby(["categoria", "producto"])["cantidad"]
        .sum()
        .reset_index()
        .sort_values("cantidad", ascending=False)
    )
    if prod_total.empty:
        st.info("Sin datos.")
    else:
        col_a, col_b = st.columns([2, 1])
        with col_a:
            fig3 = px.treemap(
                prod_total,
                path=["categoria", "producto"],
                values="cantidad",
                color="cantidad",
                color_continuous_scale="Blues",
                template="plotly_white",
                height=380,
            )
            fig3.update_layout(margin=dict(t=10))
            st.plotly_chart(fig3, use_container_width=True)
        with col_b:
            fig4 = px.bar(
                prod_total.head(12),
                x="cantidad", y="producto",
                orientation="h",
                color="categoria",
                labels={"cantidad": "Total", "producto": ""},
                template="plotly_white",
                color_discrete_sequence=px.colors.qualitative.Set2,
                height=380,
            )
            fig4.update_layout(
                showlegend=False,
                margin=dict(t=10, l=10),
                yaxis=dict(categoryorder="total ascending"),
            )
            st.plotly_chart(fig4, use_container_width=True)

    # Average consumption per opening
    st.markdown('<div class="section-header">Promedio por apertura</div>', unsafe_allow_html=True)
    avg_df = (
        dff.groupby("producto")["cantidad"]
        .mean()
        .reset_index()
        .rename(columns={"cantidad": "promedio"})
        .sort_values("promedio", ascending=False)
    )
    fig5 = px.bar(
        avg_df,
        x="producto", y="promedio",
        labels={"producto": "Producto", "promedio": "Promedio por apertura"},
        template="plotly_white",
        color="promedio",
        color_continuous_scale="Teal",
        height=320,
    )
    fig5.update_layout(coloraxis_showscale=False, margin=dict(t=10, b=80),
                       xaxis=dict(tickangle=-35))
    st.plotly_chart(fig5, use_container_width=True)

# ── TAB 3: By category ────────────────────────────────────────────────────────
with tab3:
    st.markdown('<div class="section-header">Distribución por categoría</div>', unsafe_allow_html=True)
    cat_total = dff.groupby("categoria")["cantidad"].sum().reset_index()
    col_c, col_d = st.columns(2)
    with col_c:
        fig6 = px.pie(
            cat_total, names="categoria", values="cantidad",
            hole=0.4, template="plotly_white",
            color_discrete_sequence=px.colors.qualitative.Set3,
            height=340,
        )
        fig6.update_traces(textinfo="percent+label")
        fig6.update_layout(showlegend=False, margin=dict(t=10))
        st.plotly_chart(fig6, use_container_width=True)
    with col_d:
        cat_mes = (
            dff.groupby(["mes_label", "categoria"])["cantidad"]
            .sum()
            .reset_index()
            .sort_values("mes_label")
        )
        fig7 = px.area(
            cat_mes, x="mes_label", y="cantidad",
            color="categoria",
            labels={"mes_label": "Mes", "cantidad": "Unidades", "categoria": "Categoría"},
            template="plotly_white",
            color_discrete_sequence=px.colors.qualitative.Set3,
            height=340,
        )
        fig7.update_layout(margin=dict(t=10, b=60), xaxis=dict(tickangle=-30))
        st.plotly_chart(fig7, use_container_width=True)

# ── TAB 4: Raw data ───────────────────────────────────────────────────────────
with tab4:
    st.markdown('<div class="section-header">Registros filtrados</div>', unsafe_allow_html=True)
    show_df = dff[["categoria", "producto", "cantidad", "fecha_inicio", "fecha_fin", "mes_label"]].copy()
    show_df.columns = ["Categoría", "Producto", "Cantidad", "Fecha inicio", "Fecha fin", "Mes"]
    st.dataframe(show_df.sort_values("Fecha inicio"), use_container_width=True, height=400)
    csv = show_df.to_csv(index=False).encode("utf-8")
    st.download_button("⬇️ Descargar CSV", csv, "consumo_filtrado.csv", "text/csv")

st.caption("Mercado 8AM · Datos reales del hogar · Visualizado con Streamlit + Plotly")
